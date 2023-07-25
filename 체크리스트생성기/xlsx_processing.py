import openpyxl as xl
from openpyxl.styles import Font, Alignment, Color
from tqdm import tqdm

def postprocess_cashshop(xlFileName):
    wb = xl.load_workbook(xlFileName,data_only = True)
    sheetList = wb.sheetnames
    ws = wb[sheetList[0]]
    ws.column_dimensions['b'].width = 17
    ws.column_dimensions['c'].width = 17
    ws.column_dimensions['d'].width = 17
    ws.column_dimensions['e'].width = 50

    firstRow = 2
    lastRow = ws.max_row
    startRow_B =0
    startValue_B =""
    startRow_C = 0
    startRow_D = 0

    tqdmCount1 = 0
    print("엑셀 서식 처리중...")
    for i in tqdm(range(firstRow, lastRow+1)):
        tqdmCount1+=1
        #print(i)

        """서버 카테고리 합치기"""
        if (ws['b'+str(i)].value is not None) :
            if startRow_B == 0  :
                startRow_B = i
                startValue_B = ws['b'+str(i)].value
                #print(startRow_B)
            else :
                #firstTargetCell =  "C"+str(startRow_C)
                if ( ws['b'+str(i)].value != startValue_B) :
                    mergeTargetCell = "B"+str(startRow_B)+":B"+str(i-1)
                    ws.merge_cells(mergeTargetCell)
                    startValue_B = ws['b'+str(i)].value
                    startRow_B = i

        """판매 분류 카테고리 합치기"""
        if ws['c'+str(i)].value is not None:
            if startRow_C == 0 :
                startRow_C = i
                #print(startRow)
                startValue_C = ws['c'+str(i)].value
            else :
                # firstTargetCell =  "C"+str(startRow_C)
                # mergeTargetCell = "C"+str(startRow_C)+":C"+str(i-1)
                # ws.merge_cells(mergeTargetCell)
                # startRow_C = i
                if ( ws['c'+str(i)].value != startValue_C) :
                    mergeTargetCell = "c"+str(startRow_C)+":c"+str(i-1)
                    ws.merge_cells(mergeTargetCell)
                    startValue_C = ws['c'+str(i)].value
                    startRow_C = i


        if ws['d'+str(i)].value is not None:
            if startRow_D == 0 :
                startRow_D = i
                #print(startRow)
            else :
                firstTargetCell =  "D"+str(startRow_D)
                mergeTargetCell = "D"+str(startRow_D)+":D"+str(i-1)
                ws.merge_cells(mergeTargetCell)
                startRow_D = i


        ws['b'+str(i)].alignment = Alignment(
            horizontal='center'
            ,vertical='top'
            ,wrap_text=True)
        ws['b'+str(i)].font = Font(size = 9, bold = True)
        ws['c'+str(i)].alignment = Alignment(
            horizontal='center'
            ,vertical='top'
            ,wrap_text=True)
        ws['c'+str(i)].font = Font(size = 9, bold = True)
        ws['d'+str(i)].alignment = Alignment(
            horizontal='center'
            ,vertical='top'
            ,wrap_text=True)
        ws['d'+str(i)].font = Font(size = 9, bold = True)
        ws['e'+str(i)].alignment = Alignment(
            horizontal='left'
            ,vertical='top'
            ,wrap_text=True)
        ws['e'+str(i)].font = Font(size = 9, bold = False)
        
        #ws['e'+str(i)].value = process_temp_str(str(ws['e'+str(i)].value))


    #예외 마지막 셀병합
    ws.merge_cells("B"+str(startRow_B)+":B"+str(lastRow))
    #ws.merge_cells("C"+str(startRow_C)+":C"+str(lastRow))
    #ws.merge_cells("D"+str(startRow_D)+":D"+str(lastRow))

    #ws = highlight_belonging(ws)
    #ws = find_and_replace(ws,"귀속","귀속")
    #ws = highlight_star_cells(ws)
    wb.save(xlFileName)



def postprocess_cashshop(xlFileName):
    wb = xl.load_workbook(xlFileName,data_only = True)
    sheetList = wb.sheetnames
    ws = wb[sheetList[0]]
    ws.column_dimensions['a'].width = 17
    ws.column_dimensions['b'].width = 17
    #ws.column_dimensions['d'].width = 17
    ws.column_dimensions['c'].width = 50

    firstRow = 2
    lastRow = ws.max_row
    startRow_B =0
    startValue_B =""
    startRow_C = 0
    startRow_D = 0

    tqdmCount1 = 0
    print("엑셀 서식 처리중...")
    for i in tqdm(range(firstRow, lastRow+1)):
        tqdmCount1+=1
        #print(i)

        """서버 카테고리 합치기"""
        if (ws['a'+str(i)].value is not None) :
            if startRow_B == 0  :
                startRow_B = i
                startValue_B = ws['a'+str(i)].value
                #print(startRow_B)
            else :
                #firstTargetCell =  "C"+str(startRow_C)
                if ( ws['a'+str(i)].value != startValue_B) :
                    mergeTargetCell = "a"+str(startRow_B)+":a"+str(i-1)
                    ws.merge_cells(mergeTargetCell)
                    startValue_B = ws['a'+str(i)].value
                    startRow_B = i

        """판매 분류 카테고리 합치기"""
        if ws['b'+str(i)].value is not None:
            if startRow_C == 0 :
                startRow_C = i
                startValue_C = ws['b'+str(i)].value
            else :
                if ( ws['b'+str(i)].value != startValue_C) :
                    mergeTargetCell = "b"+str(startRow_C)+":b"+str(i-1)
                    ws.merge_cells(mergeTargetCell)
                    startValue_C = ws['b'+str(i)].value
                    startRow_C = i


        # if ws['d'+str(i)].value is not None:
        #     if startRow_D == 0 :
        #         startRow_D = i
        #     else :
        #         firstTargetCell =  "D"+str(startRow_D)
        #         mergeTargetCell = "D"+str(startRow_D)+":D"+str(i-1)
        #         ws.merge_cells(mergeTargetCell)
        #         startRow_D = i


        ws['a'+str(i)].alignment = Alignment(
            horizontal='center'
            ,vertical='top'
            ,wrap_text=True)
        ws['a'+str(i)].font = Font(size = 9, bold = True)
        ws['b'+str(i)].alignment = Alignment(
            horizontal='center'
            ,vertical='top'
            ,wrap_text=True)
        ws['b'+str(i)].font = Font(size = 9, bold = True)

        # ws['c'+str(i)].alignment = Alignment(
        #     horizontal='center'
        #     ,vertical='top'
        #     ,wrap_text=True)
        # ws['c'+str(i)].font = Font(size = 9, bold = True)

        # ws['d'+str(i)].alignment = Alignment(
        #     horizontal='center'
        #     ,vertical='top'
        #     ,wrap_text=True)
        # ws['d'+str(i)].font = Font(size = 9, bold = True)

        ws['c'+str(i)].alignment = Alignment(
            horizontal='left'
            ,vertical='top'
            ,wrap_text=True)
        ws['c'+str(i)].font = Font(size = 9, bold = False)
        
        #ws['e'+str(i)].value = process_temp_str(str(ws['e'+str(i)].value))


    #예외 마지막 셀병합
    ws.merge_cells("A"+str(startRow_B)+":A"+str(lastRow))
    #ws.merge_cells("B"+str(startRow_B)+":B"+str(lastRow))
    #ws.merge_cells("C"+str(startRow_C)+":C"+str(lastRow))
    #ws.merge_cells("D"+str(startRow_D)+":D"+str(lastRow))

    #ws = highlight_belonging(ws)
    #ws = find_and_replace(ws,"귀속","귀속")
    #ws = highlight_star_cells(ws)
    wb.save(xlFileName)

