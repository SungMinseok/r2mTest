import re
import csv
import time
from openpyxl import load_workbook
import os
import pandas as pd

output_file_name = f"로그정리_{time.strftime('%y%m%d_%H%M%S')}.xlsx"

def 로그정리(input_file_name):
    if input_file_name == "뽑기 확률테스트" :

        # 입력 및 출력 파일 이름 설정
        #output_file_name = f"{input_file_name} 정리.csv"

        # CSV 파일에 저장할 헤더 설정
        csv_header = ["뽑기ID","그룹ID", "ID", "보상선택횟수", "확률"]

        # 정규 표현식 패턴 설정
        pattern_title = re.compile(r"뽑기번호 : \((\d+)\)")
        pattern_title_1 = re.compile(r"그룹번호 \((\d+)\)")
        pattern_config = re.compile(r"ID : \((\d+)\), .* 보상선택횟수 : \((\d+)\), 확률 : \(([\d.]+)\)")

        # 결과를 저장할 리스트 초기화
        results = []

        # 파일 읽기 및 데이터 추출        
        with open(f'{input_file_name}.txt', "r", encoding="utf-16") as file:

            lines = file.readlines()

        current_title = None
        current_title_1 = None
        for line in lines:
            title_match = pattern_title.search(line)
            title_match_1 = pattern_title_1.search(line)
            config_match = pattern_config.search(line)
            
            if title_match:
                current_title = title_match.group(1)
            elif title_match_1:
                current_title_1 = title_match_1.group(1)
            elif config_match:
                id_value = config_match.group(1)
                select_count = config_match.group(2)
                probability = config_match.group(3)
                results.append([current_title, current_title_1, id_value, select_count, probability])

        # # 추출한 데이터를 CSV 파일로 저장
        # with open(output_file_name, "w", newline="", encoding="utf-16") as csv_file:
        #     csv_writer = csv.writer(csv_file,delimiter='\t')
        #     csv_writer.writerow(csv_header)
        #     csv_writer.writerows(results)

        
        # 데이터프레임 생성
        df = pd.DataFrame(results, columns=csv_header)
        
        try:
            book = load_workbook(output_file_name)
            writer = pd.ExcelWriter(output_file_name, engine='openpyxl')
            writer.book = book
            df.to_excel(writer, sheet_name=input_file_name, index=False)
            writer.save()
        except FileNotFoundError:
            df.to_excel(output_file_name, sheet_name=input_file_name, index=False)

        print(f"{output_file_name} 파일로 저장되었습니다.")


    elif input_file_name == "다시뽑기 확률테스트" :

        # 입력 및 출력 파일 이름 설정
        #output_file_name = f"{input_file_name} 정리.csv"

        # CSV 파일에 저장할 헤더 설정
        csv_header = ["획득경로","종류","원본ID","그룹ID", "ID", "보상선택횟수", "확률"]

        # 정규 표현식 패턴 설정
        #pattern_title = re.compile(r'\[---- (.*?) 획득 (.*?) 다시뽑기')
        pattern_title = re.compile(r'\[---- (.*?)획득')
        pattern_title_1 = re.compile(r'\획득 (.*?) 다시뽑기')
        pattern_title_2 = re.compile(r"원본카드 ID \((\d+)\)")
        pattern_title_3 = re.compile(r"그룹번호 \((\d+)\)")
        pattern_config = re.compile(r"ID : \((\d+)\), .* 보상선택횟수 : \((\d+)\), 확률 : \(([\d.]+)\)")

        # 결과를 저장할 리스트 초기화
        results = []

        # 파일 읽기 및 데이터 추출
        with open(f'{input_file_name}.txt', "r", encoding="utf-16") as file:
            lines = file.readlines()

        current_title = None
        current_title_1 = None
        current_title_2 = None
        current_title_3 = None
        for line in lines:
            title_match = pattern_title.search(line)
            title_match_1 = pattern_title_1.search(line)
            title_match_2 = pattern_title_2.search(line)
            title_match_3 = pattern_title_3.search(line)
            config_match = pattern_config.search(line)
            
            if title_match:
                current_title = title_match.group(1)
            if title_match_1:
                current_title_1 = title_match_1.group(1)
            if title_match_2:
                current_title_2 = title_match_2.group(1)
            if title_match_3:
                current_title_3 = title_match_3.group(1)
            elif config_match:
                id_value = config_match.group(1)
                select_count = config_match.group(2)
                probability = config_match.group(3)
                results.append([current_title, current_title_1, current_title_2, current_title_3, id_value, select_count, probability])

        # # 추출한 데이터를 CSV 파일로 저장
        # with open(output_file_name, "w", newline="", encoding="utf-16") as csv_file:
        #     csv_writer = csv.writer(csv_file,delimiter='\t')
        #     csv_writer.writerow(csv_header)
        #     csv_writer.writerows(results)


        
        # 데이터프레임 생성
        df = pd.DataFrame(results, columns=csv_header)
        
        try:
            book = load_workbook(output_file_name)
            writer = pd.ExcelWriter(output_file_name, engine='openpyxl')
            writer.book = book
            df.to_excel(writer, sheet_name=input_file_name, index=False)
            writer.save()
        except FileNotFoundError:
            df.to_excel(output_file_name, sheet_name=input_file_name, index=False)


        print(f"{output_file_name} 파일로 저장되었습니다.")

로그정리("뽑기 확률테스트")
로그정리("다시뽑기 확률테스트")