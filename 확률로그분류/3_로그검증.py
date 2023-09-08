# import pandas as pd
# import csv
# from openpyxl import load_workbook

# # 1. '뽑기_구성리스트.csv' 파일을 읽어와 데이터프레임으로 저장
# csv_file_name = "뽑기_구성리스트.csv"
# df = pd.read_csv(csv_file_name, delimiter='\t',encoding='utf-16')

# # 2. '실데이터.xlsx' 파일을 읽어온 후 '뽑기 변신카드 보상 정보' 시트 데이터를 불러와 딕셔너리로 저장
# excel_file_name = "실데이터.xlsx"
# design_data = pd.read_excel(excel_file_name, sheet_name="뽑기 변신카드 보상 정보")
# design_dict = design_data.groupby(['mGroupID', 'mTransformCardID'])['mRate'].max().to_dict()

# # 3. '요구리스트.txt' 파일을 읽어 ID와 Desc 정보를 추출하고 '뽑기검증.xlsx' 파일에 시트로 추가
# with open('요구리스트.txt', 'r', encoding='utf-8') as file:
#     lines = file.readlines()

# for line in lines:
#     id_, desc = line.strip().split(',')
#     id_ = int(id_)
#     if id_ in df['뽑기ID'].values:
#         temp = df[df['뽑기ID'] == int(id_)]
#         temp['기획확률'] = temp.apply(lambda row: design_dict.get((row['뽑기ID'], int(row['ID'])), ''), axis=1)
#         temp = temp.drop(columns=['뽑기ID']).rename(columns={'ID': '아이디', '보상선택횟수': '선택횟수', '확률': '보상확률'})

#         # '뽑기검증.xlsx' 파일에 시트로 추가
#         with pd.ExcelWriter('뽑기검증.xlsx', mode='a', engine='openpyxl') as writer:
#             temp.to_excel(writer, sheet_name=desc, index=False)

# print("'뽑기검증.xlsx'에 데이터를 추가했습니다.")


import pandas as pd
import csv
from openpyxl import load_workbook
from openpyxl import Workbook
import os
import time
import matplotlib.pyplot as plt
plt.rc('font', family='Malgun Gothic') # For Windows

로그정리문서 = '로그정리_230908_141015.xlsx'

resultBasicDir = f"./result"
if not os.path.isdir(resultBasicDir) :
    os.mkdir(resultBasicDir)
    
result_file_name = f"./result/로그검증_{time.strftime('%y%m%d_%H%M%S')}.xlsx"

# 1. '요구리스트.txt' 파일을 읽어 ID와 Desc 정보를 추출
with open('요구리스트.txt', 'r', encoding='utf-8') as file:
    lines = file.readlines()

# # 2. '뽑기_구성리스트.csv' 파일을 읽어와 데이터프레임으로 저장
# csv_file_name = "뽑기_구성리스트.csv"
# df = pd.read_csv(csv_file_name, delimiter='\t',encoding='utf-16')
# try:
#     _sheet_name = "뽑기 확률테스트"
#     로그정리_뽑기 = pd.read_excel(로그정리문서, sheet_name=_sheet_name)
# except pd.errors.SheetNotFound:
#     print(f'로그정리문서 내 시트 없음 : {_sheet_name}')
# try:
#     _sheet_name = "다시뽑기 확률테스트"
#     로그정리_다시뽑기 = pd.read_excel(로그정리문서, sheet_name=_sheet_name)
# except pd.errors.SheetNotFound:
#     print(f'로그정리문서 내 시트 없음 : {_sheet_name}')
# try:
#     _sheet_name = "합성 확률테스트"
#     로그정리_합성 = pd.read_excel(로그정리문서, sheet_name=_sheet_name)
# except pd.errors.SheetNotFound:
#     print(f'로그정리문서 내 시트 없음 : {_sheet_name}')

sheet_names = ["뽑기 확률테스트",
                "다시뽑기 확률테스트", 
                "합성 확률테스트"
                ]

로그정리문서모음 = {}

for _sheet_name in sheet_names:
    try:
        df = pd.read_excel(로그정리문서, sheet_name=_sheet_name)
        # 시트명을 키로 사용하여 DataFrame을 저장
        로그정리문서모음[_sheet_name] = df
    except Exception as e:
        print(f'{e} 로그정리문서 내 시트 없음 : {_sheet_name}')

# 각 DataFrame 변수에 접근 가능
# print(dfs["뽑기 확률테스트"])
# print(dfs["다시뽑기 확률테스트"])
# print(dfs["합성 확률테스트"])

















# 3. '실데이터.xlsx' 파일을 읽어온 후 '뽑기 변신카드 보상 정보' 시트 데이터를 불러와 딕셔너리로 저장
excel_file_name = "실데이터.xlsx"
#임시기획 = pd.read_excel(excel_file_name, sheet_name="뽑기 변신카드 보상 정보")
실데이터모음 = {}
실데이터모음["뽑기 변신카드 보상 정보"] = pd.read_excel(excel_file_name, sheet_name="뽑기 변신카드 보상 정보").groupby(['mGroupID', 'mTransformCardID'])['mRate'].max().to_dict()
실데이터모음["뽑기 서번트카드 보상 정보"] = pd.read_excel(excel_file_name, sheet_name="뽑기 서번트카드 보상 정보").groupby(['mGroupID', 'mServantCardID'])['mRate'].max().to_dict()
실데이터모음["변신 카드 조합"] = pd.read_excel(excel_file_name, sheet_name="변신 카드 조합").groupby(['mRarity', 'mResult'])['mRatio'].max().to_dict()
실데이터모음["서번트 카드 조합"] = pd.read_excel(excel_file_name, sheet_name="서번트 카드 조합").groupby(['mRarity', 'mResult'])['mRatio'].max().to_dict()


for line in lines:
    new_line = line.strip().split(',')
    prob_id = new_line[0]
    other_line = ','.join(new_line[1:]) if len(new_line) > 1 else ""

    
    if prob_id == '1' : #뽑기
        로그 = 로그정리문서모음["뽑기 확률테스트"]

        id_, group_id, type, desc = other_line.strip().split(',')
        id_ = int(id_)
        if id_ in 로그['뽑기ID'].values:
            temp = 로그[로그['뽑기ID'] == int(id_)]
            if type == "변신":
                design_dict = 실데이터모음["뽑기 변신카드 보상 정보"]
            elif type == "서번트":
                design_dict = 실데이터모음["뽑기 서번트카드 보상 정보"]
            
            temp['기획확률(%)'] = temp.apply(lambda row: design_dict.get((row['그룹ID'], int(row['ID'])), ''), axis=1)
            temp['오차(%)'] = round((temp['기획확률(%)']-temp['확률'])/temp['기획확률(%)'],4)
            temp = temp.rename(columns={'ID': '카드ID', '보상선택횟수': '선택횟수', '확률': '인게임확률(%)'})

            plt.figure(figsize=(8, 6))
            temp.boxplot(column='오차(%)')
            plt.title(f'{desc} 오차')
            plt.ylabel('오차(%)')
            plt.show()
            # '뽑기검증.xlsx' 파일에 시트로 추가 또는 새로 생성
            try:
                book = load_workbook(result_file_name)
                writer = pd.ExcelWriter(result_file_name, engine='openpyxl')
                writer.book = book
                temp.to_excel(writer, sheet_name=desc, index=False)
                writer.save()
            except FileNotFoundError:
                temp.to_excel(result_file_name, sheet_name=desc, index=False)
    
    # elif prob_id == '2' or prob_id == '3': #합성
    #     로그 = 로그정리문서모음["합성 확률테스트"]

    #     rarity, desc = other_line.strip().split(',')
    #     rarity = int(rarity)
    #     if group_id in 로그['그룹ID'].values:
    #         temp = 로그[(로그['원본ID'] == int(origin_id))&(로그['그룹ID'] == int(group_id))]
    #         if prob_id == "2" :
    #             design_dict = 실데이터모음["변신 카드 조합"]            
    #         if prob_id == "3"":
    #             design_dict = 실데이터모음["서번트 카드 조합"]
    #         try :
    #             temp['기획확률(%)'] = temp.apply(lambda row: design_dict.get((row['그룹ID'], int(row['ID'])), ''), axis=1)
    #             원본ID확률 = design_dict.get((group_id,origin_id),'')
    #             #print(원본ID확률)
    #             temp['기획확률(%)'] =round(temp['기획확률(%)']/temp['기획확률(%)'].sum()*100,4)#round(최종결과표['확률']  * 100 / 대상제외총확률 ,4)
    #             temp['오차(%)'] = round((temp['기획확률(%)']-temp['확률'])/temp['기획확률(%)'],4)
    #             temp = temp.rename(columns={'ID': '카드ID', '보상선택횟수': '선택횟수', '확률': '인게임확률(%)'})
    #         except:
    #             pass

    #         # '뽑기검증.xlsx' 파일에 시트로 추가 또는 새로 생성
    #         try:
    #             book = load_workbook(result_file_name)
    #             writer = pd.ExcelWriter(result_file_name, engine='openpyxl')
    #             writer.book = book
    #             temp.to_excel(writer, sheet_name=desc, index=False)
    #             writer.save()
    #         except FileNotFoundError:
    #             temp.to_excel(result_file_name, sheet_name=desc, index=False)


    elif prob_id == '14' or prob_id == '16': #다시뽑기
        로그 = 로그정리문서모음["다시뽑기 확률테스트"]

        origin_id, group_id, desc = other_line.strip().split(',')
        origin_id = int(origin_id)
        group_id = int(group_id)
        if group_id in 로그['그룹ID'].values:
            temp = 로그[(로그['원본ID'] == int(origin_id))&(로그['그룹ID'] == int(group_id))]
            if prob_id == "14" or prob_id == "15":
                design_dict = 실데이터모음["뽑기 변신카드 보상 정보"]            
            if prob_id == "16" or prob_id == "17":
                design_dict = 실데이터모음["뽑기 서번트카드 보상 정보"]
            try :
                temp['기획확률(%)'] = temp.apply(lambda row: design_dict.get((row['그룹ID'], int(row['ID'])), ''), axis=1)
                원본ID확률 = design_dict.get((group_id,origin_id),'')
                #print(원본ID확률)
                temp['기획확률(%)'] =round(temp['기획확률(%)']/temp['기획확률(%)'].sum()*100,4)#round(최종결과표['확률']  * 100 / 대상제외총확률 ,4)
                temp['오차(%)'] = round((temp['기획확률(%)']-temp['확률'])/temp['기획확률(%)'],4)
                temp = temp.rename(columns={'ID': '카드ID', '보상선택횟수': '선택횟수', '확률': '인게임확률(%)'})
            except:
                pass

            # '뽑기검증.xlsx' 파일에 시트로 추가 또는 새로 생성
            try:
                book = load_workbook(result_file_name)
                writer = pd.ExcelWriter(result_file_name, engine='openpyxl')
                writer.book = book
                temp.to_excel(writer, sheet_name=desc, index=False)
                writer.save()
            except FileNotFoundError:
                temp.to_excel(result_file_name, sheet_name=desc, index=False)

print("'뽑기검증.xlsx'에 데이터를 추가했습니다.")
os.startfile(os.path.normpath(result_file_name))
